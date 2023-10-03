import argparse
import requests
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

API_BASE_URL = "http://localhost:8000"


def get_row_color(hu):
    today = datetime.now().date()
    three_months = today - timedelta(days=90)
    twelve_months = today - timedelta(days=365)

    hu = datetime.strptime(hu, "%Y-%m-%d").date()

    if hu >= three_months:
        return "007500"  # Green
    
    elif hu >= twelve_months:
        return "FFA500"  # Orange

    else:
        return "b30000"  # Red

def generate_excel_file(csv_filename, keys, colored):

    with open(csv_filename, "r") as file:
        files = {'csv_file': (csv_filename, file)}
        response = requests.post(API_BASE_URL+"/process-csv/", files=files, data={'keys': keys, 'colored': colored})
        file.close()

    
    response_df = pd.DataFrame(response.json())

    # Sort rows by gruppe values
    response_df = response_df.sort_values(by='gruppe', ascending=True)

    # Add rnr to keys, if rnr is already given move it to be the first coloumn
    if "rnr" in keys:
        keys.pop(keys.index("rnr"))
    keys.insert(0, "rnr")

    color_codes = {}
    hu_values = {}
    
    # Get the index of LabelIds coloumn if given
    if "labelIds" in keys:
        labelIds_coloumn_index = keys.index("labelIds") + 1
    
    # Store the ColorCodes for labelIds and the hu values of rows using dictionaries, before dropping these coloumns if not given in keys
    for index, data in response_df.iterrows():
        if data["labelIds"]:
            color_codes[data["labelIds"]] = "000000" if not data["colorCode"] else data["colorCode"][1:]
        if colored:
            hu_values[data["rnr"]] = data["hu"]
        
    # Dropping unwanted coloumns that are not given (except rnr)
    columns_to_delete = [col for col in response_df.columns if col not in keys]
    response_df.drop(columns=columns_to_delete, inplace=True)
    response_df = response_df[keys]

    # Create Sheet object and with thin border object
    wb = Workbook()
    ws = wb.active
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Creating the coloumns for the spreadsheet with styling
    for col_idx, column in enumerate(response_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        max_length = max(response_df[column].astype(str).apply(len))
        adjusted_width =  min((max_length + 2) * 1.2, 100) # Adding a little extra width for padding
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Filling the spreadsheet and coloring the rows accordingly if colored flag is set to True
    for r_idx, row in enumerate(response_df.itertuples(), start=2):
        if colored:
            row_color = get_row_color(hu_values[row[1]])
        for c_idx, value in enumerate(row[1:], start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            if colored:
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            if "labelIds" in keys and c_idx == labelIds_coloumn_index and row[c_idx]:
                cell.font = Font(color=color_codes.get(row[c_idx], "000000"))

    # Creating the excel file
    current_date_iso_formatted = datetime.now().date().isoformat()
    excel_filename = f"vehicles-{current_date_iso_formatted}.xlsx"
    wb.save(excel_filename)

    print(f"vehicles-{current_date_iso_formatted}.xlsx was created successfully.")


if __name__ == '__main__':
    # Creating argument parser
    parser = argparse.ArgumentParser(description='Transmit CSV to a REST API and generate an Excel file.')

    # Adding the -k/--keys argument (list of strings)
    parser.add_argument('-k', '--keys', nargs='+', help='List of keys to include.')

    # Adding the -c/--colored argument (boolean flag)
    parser.add_argument('-c', '--colored', default=True, help='Enable row coloring.')

    # Parsing the command-line arguments
    args = parser.parse_args()

    csv_filename = 'vehicles.csv'

    if args.colored == "True":
        colored = True
    elif args.colored == "False":
        colored = False
    else:
        print("Invalid value for flag -c/--colored")

    generate_excel_file(csv_filename, args.keys, colored)
